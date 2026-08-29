#!/usr/bin/env python3
"""Create the review-oriented canonical JSON export from the private workbook."""
from pathlib import Path
import hashlib, json

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT.parent / "catechism-app" / "Catechism Canonical Validation.xlsx"
EXPORT = ROOT / "content" / "catechism.canonical.json"

def records(sheet):
    rows = list(sheet.iter_rows(values_only=True)); headers = rows[3]
    return [dict(zip(headers, row)) for row in rows[4:] if row[0]]

def build(source=SOURCE):
    from openpyxl import load_workbook
    wb = load_workbook(source, data_only=True, read_only=True)
    questions = records(wb["Questions"]); passages = {r["Passage ID"]:r for r in records(wb["Scripture Passages"])}; maps = {r["Question ID"]:r for r in records(wb["Question Scripture Map"])}; teaching = {r["Question ID"]:r for r in records(wb["Teaching Notes"])}
    assert len(questions) == len(passages) == len(maps) == 62
    out=[]
    for q in questions:
        qid=q["Question ID"]; mapping=maps[qid]; passage=passages[mapping["Passage ID"]]; note=teaching[qid]
        assert q["Canonical Full Answer"] and passage["Exact ESV Text"]
        out.append({"id":qid,"number":q["Number"],"question":q["Canonical Question"],"answer":q["Canonical Full Answer"],"bookletPage":q["Booklet Page"],"questionSourceStatus":q["Source Check Status"],"questionApprovalStatus":q["Human Approval Status"],"scripture":{"reference":passage["Exact ESV Reference"],"text":passage["Exact ESV Text"].strip(),"verificationStatus":passage["ESV Verification Status"],"referenceReviewStatus":mapping["Reference Review Status"],"theologicalReviewStatus":mapping["Theological Review Status"],"authorizedSource":passage["Authorized ESV Source"],"note":mapping["Notes"] or ""},"teaching":{"status":note["Teaching Review Status"],"meaning":note["Plain-Language Meaning"] or "","connection":note["Connection to Scripture"] or "","whyItMatters":note["Why It Matters"] or ""}})
    notice=next(v for row in wb["README"].iter_rows(values_only=True) for v in row if isinstance(v,str) and v.startswith("Scripture quotations are from the ESV® Bible"))
    return {"schemaVersion":1,"sourceWorkbookSha256":hashlib.sha256(source.read_bytes()).hexdigest(),"esvStandardNotice":notice,"questions":out}

def render(source=SOURCE): return json.dumps(build(source),ensure_ascii=False,indent=2)+"\n"
def main(): EXPORT.parent.mkdir(parents=True,exist_ok=True); EXPORT.write_text(render(),encoding="utf-8"); print(f"Exported 62 canonical records to {EXPORT}")
if __name__=="__main__": main()
